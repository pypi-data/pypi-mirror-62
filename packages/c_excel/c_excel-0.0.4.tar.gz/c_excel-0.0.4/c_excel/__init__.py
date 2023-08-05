import xlwt
import openpyxl
import xlrd
import os
from xlutils.copy import copy


class Excel(object):
    def __init__(self,file_name,excel_rev="03"):
        self.excel_rev=excel_rev
        self.file_name=file_name

    def write_cell(self,sheet,row,col,value):
        if self.excel_rev=='03':
            sheet.write(row-1,col-1,value)
        else:
            sheet.cell(row,col,value)

    def open_excel(self):
        workbook= xlrd.open_workbook(self.file_name)
        return workbook

    def create_excel(self):
        if self.excel_rev=="03":
            workbook = xlwt.Workbook()
        else:
            workbook = openpyxl.Workbook()
        return workbook


    def create_sheet(self,workbook,sheet_name):
        if self.excel_rev=="03":
            # workbook = xlwt.Workbook()
            sheet=workbook.add_sheet(sheet_name)
        else:
            # workbook = openpyxl.Workbook()
            sheet = workbook.create_sheet(sheet_name)
        return sheet

    def remove_sheet(self,workbook,sheet_name):
        workbook._Workbook__worksheets = [ worksheet for worksheet in workbook._Workbook__worksheets if worksheet.name != sheet_name ]
        return workbook

    def remove_excel_sheet(self,sheet_name):
        if os.path.exists(self.file_name):
            workbook_old = self.open_excel()
            workbook = copy(workbook_old)
            if sheet_name in workbook_old.sheet_names():
                workbook=self.remove_sheet(workbook,sheet_name)
                self.save_file(workbook)
        else:
            raise("excel file not exists!")
 

    def add_sheet(self,sheet_name):
        if os.path.exists(self.file_name):
            workbook_old = self.open_excel()
            workbook = copy(workbook_old)
            if sheet_name in workbook_old.sheet_names():
                workbook=self.remove_sheet(workbook,sheet_name)
                self.save_file(workbook)
                workbook_old = self.open_excel()
                workbook = copy(workbook_old)
        else:
            workbook=self.create_excel()
        sheet=workbook.add_sheet(sheet_name)
        return workbook,sheet


    def save_file(self,workbook):
        if self.excel_rev=="03":
            workbook.save(self.file_name)
        else:
            workbook.save_file(self.file_name)        


