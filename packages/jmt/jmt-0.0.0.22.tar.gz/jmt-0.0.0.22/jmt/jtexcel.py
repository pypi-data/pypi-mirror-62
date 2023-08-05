# -*- coding: utf-8 -*-
"""
Created on Sun Jan 28 13:23:16 2018

@author: jaredmt
"""

'''============excel functions==========='''        
        
def excelToMatrix(file_name,sheet_name):
    '''convert all excell data into a matrix
    for test: use file_name = "excelTesting.xlsx" and 
    sheet_name = "Sheet1"
    note: this can only handle numerical data, no strings
    things to add: detect non numerical values or make a list of the values
    note: need to exit excel file before returning a
    alternative: read_excel (from pandas module)
    '''
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_name, read_only=True)
    ws = wb[sheet_name]

    L = list(ws.rows)   
    a=mxnlist(len(L),len(L[0]))#create LxL[0] list
    for i in range(0,len(L)):#copy L.value into list a
        for j in range(0,len(L[0])):
            a[i][j] = L[i][j].value
    wb._archive.close()
    return a
    
def matrixToExcel(file_name,sheet_name,M):
    '''writing a matrix (nested list) to excel
    alternative: ExcelWriter (from pandas module)
    '''
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_name)
    ws=wb[sheet_name]
    for i in range(0,len(M)):
        for j in range(0,len(M[0])):
            s=ws.cell(row=i+1,column=j+1)#first cell is [1][1]
            s.value = M[i][j]
    wb.save(file_name)
    wb.close()
    
def closeExcel(file_name):
    '''simply closes the excel book
    closeExcel(file_name)
    file_name = string (path+filename)
    note this does solve the 'open by another user' issue
    if the issue persists after calling this function then restart spyder
    '''
    from openpyxl import load_workbook
    wb=load_workbook(file_name,read_only=False)#can't be readonly,need to save it
    wb.save(file_name)#save over current version (overwriting)
    try: wb._archive.close()#Best way to close but doesn't always work
    except: wb.close()