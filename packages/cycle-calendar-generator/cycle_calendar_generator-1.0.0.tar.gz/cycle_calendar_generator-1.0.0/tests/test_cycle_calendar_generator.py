#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Tests for `cycle_calendar_generator` package."""


import unittest
from unittest import mock
import os
import datetime

from pyfakefs import fake_filesystem_unittest
import openpyxl
from ics import Calendar, Event
import arrow

from cycle_calendar_generator import cycle_calendar_generator

def convert_time_to_day_fraction(hours, minutes):
    input_time=datetime.datetime(1, 1, 1, hours, minutes)
    timediff = input_time-datetime.datetime.min
    secs_in_day=datetime.timedelta(days=1).total_seconds()
    return timediff.total_seconds()/secs_in_day

def convert_date_string_to_datetime(datestring):
    parsed_datestring = arrow.get(datestring, "MMMM D, YYYY")
    return parsed_datestring.datetime

def convert_timestring_to_time(timestring):
    parsed_timestring = datetime.datetime.strptime(timestring, "%I:%M %p")
    return parsed_timestring.time()

def make_setup_excel(periodTiming, cycleDaysList, yearlySchedule):
    return_value = openpyxl.Workbook()
    ws_periodTiming = return_value.create_sheet(
        cycle_calendar_generator.SHEET_SETUP_PERIOD_TIMING)
    ws_cycleDaysList = return_value.create_sheet(
        cycle_calendar_generator.SHEET_SETUP_CYCLEDAYSLIST)
    ws_yearlySchedule = return_value.create_sheet(
        cycle_calendar_generator.SHEET_SETUP_YEARLYSCHEDULE)
    for line in periodTiming:
        ws_periodTiming.append(line)
    ws_cycleDaysList.append(cycleDaysList)
    for line in yearlySchedule:
        ws_yearlySchedule.append(line)
    return return_value

def make_setup_excel_good():
    good_periodTiming = [["Period Number", "Start Time", "End Time"],
                         ["1", datetime.time(8,0), datetime.time(9,0)],
                         ["2", datetime.time(9,0), datetime.time(10,0)],
                         ["3", datetime.time(10,0), datetime.time(11,0)],
                         ["4", datetime.time(11,0), datetime.time(12,0)],
                         ["5", datetime.time(12,0), datetime.time(13,0)]]
    good_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
    good_yearlySchedule = [["Date", "Cycle Day"],
                       [convert_date_string_to_datetime("August 31, 2018"),
                        good_cycleDaysList[0]],
                       [convert_date_string_to_datetime("September 3, 2018"),
                        good_cycleDaysList[1]],
                       [convert_date_string_to_datetime("September 4, 2018"),
                        good_cycleDaysList[2]],
                       [convert_date_string_to_datetime("September 5, 2018"),
                        good_cycleDaysList[3]],
                       [convert_date_string_to_datetime("September 6, 2018"),
                        good_cycleDaysList[4]],
                       [convert_date_string_to_datetime("September 7, 2018"),
                        good_cycleDaysList[5]]]
    return make_setup_excel(good_periodTiming, good_cycleDaysList,
                            good_yearlySchedule)

class Test_get_args(unittest.TestCase):
    # Get directory from args
    """Tests function to get folder argument and give default if none given"""

    def setUp(self):
        """Set up test fixtures, if any."""

    def tearDown(self):
        """Tear down test fixtures, if any."""

    @mock.patch('.'.join(['cycle_calendar_generator',
                         'cycle_calendar_generator',
                         'argparse',
                         'ArgumentParser',
                         'parse_args']))
    def test_if_arg_is_string(self, mock_parse_args):
        """Normal input on command line"""
        mock_parse_args.return_value = \
            cycle_calendar_generator.argparse.Namespace(folder='string')
        self.assertIsInstance(cycle_calendar_generator.getArgs(), str)

    @mock.patch('.'.join(['cycle_calendar_generator',
                         'cycle_calendar_generator',
                         'argparse',
                         'ArgumentParser',
                         'parse_args']))
    @mock.patch('.'.join(['cycle_calendar_generator',
                         'cycle_calendar_generator',
                         'os',
                         'getcwd']))
    def test_gets_current_dir_if_no_arg_given(self, mock_getcwd,
                                              mock_parse_args):
        """Use current working directory if no folder given"""
        cwd_string = 'current working directory'
        mock_parse_args.return_value = \
            cycle_calendar_generator.argparse.Namespace(folder=None)
        mock_getcwd.return_value = cwd_string
        self.assertEqual(cycle_calendar_generator.getArgs(), cwd_string)

class Test_read_schedule_setup_file(fake_filesystem_unittest.TestCase):
    """Tests function to open and read schedule setup Excel file"""
    """Reading is assumed to work correctly"""
    """Tests only check that exceptions are raised properly on bad input"""
    test_folder_path = '/test-folder'
    setup_filepath = "{}/{}".format(
        test_folder_path, cycle_calendar_generator.SCHEDULE_SETUP_FILENAME)
    data_justATextDoc = "This isn't actually an Excel file."

    def setUp(self):
        self.setUpPyfakefs()
        os.mkdir(self.test_folder_path)

    def test_raises_valueerror_if_invalid_path(self):
        """If input string is not a valid folder path, raise ValueError"""
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_INVALID_FOLDER,
            cycle_calendar_generator.readScheduleSetupFile,
            '/test-notafolder')

    def test_raises_valueerror_if_no_setup_file(self):
        """If no Excel file matching preset filename exists, raise ValueError"""
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_MISSING_SETUP_FILE,
            cycle_calendar_generator.readScheduleSetupFile,
            self.test_folder_path)

    def test_raises_valueerror_if_setup_file_not_excel(self):
        """If setup file isn't an Excel file, raise ValueError"""
        # just a text file
        self.assertTrue(os.path.exists(self.test_folder_path))
        with open(self.setup_filepath, "x") as file:
            file.write(self.data_justATextDoc)
            self.assertTrue(os.path.exists(self.setup_filepath))
            self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_SETUP_FILE_NOT_EXCEL,
                cycle_calendar_generator.readScheduleSetupFile,
                self.test_folder_path)

class Test_parse_schedule_setup_file(unittest.TestCase):
    # Parse schedule setup workbook input (checking for valid data)
    # Parsing should generate:
    ## [dict]periodTiming -> [int]periodNumber: [tuple(Date, Date)](startTime, endTime)
    ## [list]CycleDaysList -> (str)cycleDay {showing all cycleDay strings}
    ## [dict]yearlySchedule -> [Date]date: [str]cycleDay
    """Tests function to parse schedule setup Excel file"""
    # Setting up good and bad Excel files
    parsed_periodTiming = {"1": (datetime.time(8, 0), datetime.time(9, 0)),
                           "2": (datetime.time(9, 0), datetime.time(10, 0)),
                           "3": (datetime.time(10, 0), datetime.time(11, 0)),
                           "4": (datetime.time(11, 0), datetime.time(12, 0)),
                           "5": (datetime.time(12, 0), datetime.time(13, 0))}
    parsed_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
    parsed_yearlySchedule = {
        datetime.date(2018, 8, 31): parsed_cycleDaysList[0],
        datetime.date(2018, 9, 3): parsed_cycleDaysList[1],
        datetime.date(2018, 9, 4): parsed_cycleDaysList[2],
        datetime.date(2018, 9, 5): parsed_cycleDaysList[3],
        datetime.date(2018, 9, 6): parsed_cycleDaysList[4],
        datetime.date(2018, 9, 7): parsed_cycleDaysList[5]}
    data_badExcel = [["This", "isn't", "the", "right"],
                     ["data", "for", "the", "parser"]]
    wb_setup_good = make_setup_excel_good()
    wb_setup_bad = openpyxl.Workbook()
    ws_bad = wb_setup_bad.active
    for line in data_badExcel:
        ws_bad.append(line)

    def test_parses_correct_setup(self):
        """Expected behavior: finds file with preset filename, opens and parses,
        returning object containing setup dicts and lists"""
        # self.assert? setup dicts and lists produced properly
        parsed_setup = cycle_calendar_generator.parseScheduleSetup(
            self.wb_setup_good)
        self.assertIsInstance(parsed_setup, cycle_calendar_generator.SetupData)
        self.assertEqual(self.parsed_periodTiming,
                         parsed_setup.periodTiming)
        self.assertEqual(self.parsed_cycleDaysList,
                         parsed_setup.cycleDaysList)
        self.assertEqual(self.parsed_yearlySchedule,
                         parsed_setup.yearlySchedule)

    def test_raises_valueerror_if_setup_unparseable(self):
        """If Excel file can't be parsed following preset format, raise ValueError"""
        # test 1: just a text file
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_INVALID_SETUP_FILE,
            cycle_calendar_generator.parseScheduleSetup,
            self.wb_setup_bad)

class Test_user_schedule_file_scanner(fake_filesystem_unittest.TestCase):
    # Check for user schedule Excel files
    # Iterate over user schedule files
    """Tests function to scan through user schedule files and generate icals"""

    # Test setup
    # TODO: Use path.join with these paths
    test_folder_path = '/test-folder'
    output_folder_path = test_folder_path + '/output'
    # Generate SetupData input object
    wb_setup = make_setup_excel_good()
    setupData = cycle_calendar_generator.parseScheduleSetup(wb_setup)

    # Generate fake files in folder
    def setUp(self):
        self.setUpPyfakefs()

    def test_raises_error_if_no_user_schedule_excel_files(self):
        # TODO: Use path.join with these paths
        folder = '/test-no-users'
        os.mkdir(folder)
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_NO_TEACHER_FILES,
            cycle_calendar_generator.userScheduleFileScanner,
            self.setupData, folder
            )

class Test_read_user_schedule_file(fake_filesystem_unittest.TestCase):
    ## Read user schedule file
    ## Check if file is Excel file (exception check)
    """Tests function to read user schedule file and generate Workbook"""
    """Reading is assumed to work correctly"""
    """Tests only check that exceptions are raised properly on bad input"""
    # TODO: Use path.join with these paths
    test_folder_path = '/test-folder'
    user_filename = 'FirstnameLastname.xlsx'
    user_filepath = "{}/{}".format(test_folder_path, user_filename)
    user_filepath_bad = "{}/{}".format('test-notafolder', user_filename)
    data_justATextDoc = "This isn't actually an Excel file."

    def setUp(self):
        self.setUpPyfakefs()
        os.mkdir(self.test_folder_path)

    def test_raises_valueerror_if_invalid_path(self):
        """If input string is not a valid folder path, raise ValueError"""
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
            cycle_calendar_generator.readTeacherScheduleFile,
            self.user_filepath_bad)

    def test_raises_valueerror_if_setup_file_not_excel(self):
        """If setup file isn't an Excel file, raise ValueError"""
        # just a text file
        self.assertTrue(os.path.exists(self.test_folder_path))
        with open(self.user_filepath, "x") as file:
            file.write(self.data_justATextDoc)
        self.assertTrue(os.path.exists(self.user_filepath))
        self.assertRaisesRegex(
            ValueError,
            cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
            cycle_calendar_generator.readTeacherScheduleFile,
            self.user_filepath)

class Test_parse_user_schedule(unittest.TestCase):
    ## Check that file's periodNumbers (in first column) match those in setup file
    ## Check that file's cycleDays (in first row) match those in setup file
    ## Iterate over cycleDay columns, generating list of objects; for each...
    ### Generate dailySchedule object, 2 properties:
    #### [str]cycleDay
    #### [dict]schedule -> [int]periodNumber: [str]className
    ### Sort list by cycleDay property
    """Tests function to parse user schedule Workbook and make data object"""
    wb_setup = make_setup_excel_good()
    setupData = cycle_calendar_generator.parseScheduleSetup(
        wb_setup
        )
    data_userSchedule = [
        ["Period Number", "A1", "B2", "C3", "D4", "E5", "F6"],
        ["1", "Grade 8", "", "", "Grade 11", "", "Grade 8"],
        ["2", "", "Grade 8", "", "", "Grade 11", ""],
        ["3", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch"],
        ["4", "Grade 11", "", "Grade 8", "", "", "Grade 11"],
        ["5", "", "Grade 11", "", "Grade 8", "", ""]]
    parsed_userSchedule = cycle_calendar_generator.ScheduleData(
        ["1", "2", "3", "4", "5"])
    parsed_userSchedule.addScheduleDay(
        ["A1", "Grade 8", "", "Lunch", "Grade 11", ""])
    parsed_userSchedule.addScheduleDay(
        ["B2", "", "Grade 8", "Lunch", "", "Grade 11"])
    parsed_userSchedule.addScheduleDay(
        ["C3", "", "", "Lunch", "Grade 8", ""])
    parsed_userSchedule.addScheduleDay(
        ["D4", "Grade 11", "", "Lunch", "", "Grade 8"])
    parsed_userSchedule.addScheduleDay(
        ["E5", "", "Grade 11", "Lunch", "", ""])
    parsed_userSchedule.addScheduleDay(
        ["F6", "Grade 8", "", "Lunch", "Grade 11", ""])
    data_badExcel = [["This", "isn't", "the", "right"],
                     ["data", "for", "the", "parser"]]
    wb_schedule_good = openpyxl.Workbook()
    sheetname_userSchedule = cycle_calendar_generator.SHEET_USER_SCHEDULE
    ws_userSchedule = wb_schedule_good.create_sheet(sheetname_userSchedule)
    for line in data_userSchedule:
        ws_userSchedule.append(line)
    wb_schedule_bad = openpyxl.Workbook()
    ws_bad = wb_schedule_bad.active
    for line in data_badExcel:
        ws_bad.append(line)

    def test_parses_correct_schedule(self):
        """Expected behavior: finds file with preset filename, opens and parses,
        returning object containing schedule dicts / lists"""
        # self.assert? schedule dicts / lists produced properly
        parsed_schedule = cycle_calendar_generator.parseTeacherSchedule(
                self.wb_schedule_good,
                self.setupData)
        self.assertIsInstance(parsed_schedule,
                              cycle_calendar_generator.ScheduleData)
        self.assertEqual(self.parsed_userSchedule.userSchedule,
                         parsed_schedule.userSchedule)

    def test_raises_valueerror_if_schedule_unparseable(self):
        """If Excel file can't be parsed following preset format, raise ValueError"""
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
                cycle_calendar_generator.parseTeacherSchedule,
                self.wb_schedule_bad, self.setupData)

    def test_raises_valueerror_if_period_numbers_dont_match(self):
        schedule_sheet = self.wb_schedule_good[
                cycle_calendar_generator.SHEET_USER_SCHEDULE]
        schedule_sheet["A6"] = "6"
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
                cycle_calendar_generator.parseTeacherSchedule,
                self.wb_schedule_good, self.setupData)

    def test_raises_valueerror_if_cycle_days_dont_match(self):
        schedule_sheet = self.wb_schedule_good[
                cycle_calendar_generator.SHEET_USER_SCHEDULE]
        schedule_sheet["G1"] = "G7"
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
                cycle_calendar_generator.parseTeacherSchedule,
                self.wb_schedule_good, self.setupData)

class Test_generate_user_schedule_calendar(unittest.TestCase):
    ## Given ScheduleData and SetupData objects, create new iCal Calendar object
    ## Iterate over date:cycleDay dict, for each...
    ### Find dailySchedule object matching cycleDay
    ### Iterate over periodNumbers, for each...
    #### Check if className exists for this periodNumber, skip this if not
    #### Generate iCal Event object
    ##### Name of event = className
    ##### Start date, end date = this date
    ##### Start time, end time = found by referencing periodTiming
    #### Append Event object to Calendar
    """Tests function to use data object to make ical object"""

    # Test input data
    # SetupData
    wb_setup = make_setup_excel_good()
    setup_data = cycle_calendar_generator.parseScheduleSetup(wb_setup)
    # ScheduleData
    data_userSchedule = [
        ["Period Number", "A1", "B2", "C3", "D4", "E5", "F6"],
        ["1", "Grade 8", "", "", "Grade 11", "", "Grade 8"],
        ["2", "", "Grade 8", "", "", "Grade 11", ""],
        ["3", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch"],
        ["4", "Grade 11", "", "Grade 8", "", "", "Grade 11"],
        ["5", "", "Grade 11", "", "Grade 8", "", ""]]
    wb_schedule_good = openpyxl.Workbook()
    ws_userSchedule = wb_schedule_good.create_sheet(
            cycle_calendar_generator.SHEET_USER_SCHEDULE)
    for line in data_userSchedule:
        ws_userSchedule.append(line)
    schedule_data_good = cycle_calendar_generator.parseTeacherSchedule(
            wb_schedule_good, setup_data)

    def test_makes_correct_ical(self):
        """Expected behavior: given SetupData and ScheduleData"""
        """creates correct Calendar object"""
        correct_calendar_data = [
                ["20180831 08:00:00", "20180831 09:00:00", "Grade 8"],
                ["20180831 10:00:00", "20180831 11:00:00", "Lunch"],
                ["20180831 11:00:00", "20180831 12:00:00", "Grade 11"],
                ["20180903 09:00:00", "20180903 10:00:00", "Grade 8"],
                ["20180903 10:00:00", "20180903 11:00:00", "Lunch"],
                ["20180903 12:00:00", "20180903 13:00:00", "Grade 11"],
                ["20180904 10:00:00", "20180904 11:00:00", "Lunch"],
                ["20180904 11:00:00", "20180904 12:00:00", "Grade 8"],
                ["20180905 08:00:00", "20180905 09:00:00", "Grade 11"],
                ["20180905 10:00:00", "20180905 11:00:00", "Lunch"],
                ["20180905 12:00:00", "20180905 13:00:00", "Grade 8"],
                ["20180906 09:00:00", "20180906 10:00:00", "Grade 11"],
                ["20180906 10:00:00", "20180906 11:00:00", "Lunch"],
                ["20180907 08:00:00", "20180907 09:00:00", "Grade 8"],
                ["20180907 10:00:00", "20180907 11:00:00", "Lunch"],
                ["20180907 11:00:00", "20180907 12:00:00", "Grade 11"]]
        correct_calendar = Calendar()
        for line in correct_calendar_data:
            e = Event()
            begin, end, name = line
            begin = arrow.get(begin, "YYYYMMDD HH:mm:ss")
            end = arrow.get(end, "YYYYMMDD HH:mm:ss")
            e.begin = begin.replace(tzinfo='local')
            e.end = end.replace(tzinfo='local')
            e.name = name
            correct_calendar.events.add(e)
        created_calendar = \
                cycle_calendar_generator.generateTeacherScheduleCalendar(
                        self.schedule_data_good, self.setup_data)
        self.assertIsInstance(created_calendar, cycle_calendar_generator.Calendar)
        created_calendar_count = len(created_calendar.events)
        correct_calendar_count = len(correct_calendar.events)
        self.assertEqual(created_calendar_count, correct_calendar_count)
        sorted_created_events = sorted(created_calendar.events)
        sorted_correct_events = sorted(correct_calendar.events)
        for i in range(created_calendar_count):
            this_created_event = sorted_created_events[i]
            this_correct_event = sorted_correct_events[i]
            self.assertEqual(this_created_event.name,
                             this_correct_event.name)
            self.assertEqual(this_created_event.begin,
                             this_correct_event.begin)
            self.assertEqual(this_created_event.end,
                             this_correct_event.end)

    def test_raises_valueerror_on_bad_yearly_schedule_cycle_day(self):
        bad_periodTiming = [
                ["Period Number", "Start Time", "End Time"],
                ["1", datetime.time(8, 0), datetime.time(9, 0)],
                ["2", datetime.time(9, 0), datetime.time(10, 0)],
                ["3", datetime.time(10, 0), datetime.time(11, 0)],
                ["4", datetime.time(11, 0), datetime.time(12, 0)],
                ["5", datetime.time(12, 0), datetime.time(13, 0)]]
        bad_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
        bad_yearlySchedule = [
                ["Date", "Cycle Day"],
                # Bad data here
                [convert_date_string_to_datetime("August 31, 2018"), "G7"],
                [convert_date_string_to_datetime("September 3, 2018"),
                 bad_cycleDaysList[1]],
                [convert_date_string_to_datetime("September 4, 2018"),
                 bad_cycleDaysList[2]],
                [convert_date_string_to_datetime("September 5, 2018"),
                 bad_cycleDaysList[3]],
                [convert_date_string_to_datetime("September 6, 2018"),
                 bad_cycleDaysList[4]],
                [convert_date_string_to_datetime("September 7, 2018"),
                 bad_cycleDaysList[5]]]
        bad_setup_excel = make_setup_excel(bad_periodTiming, bad_cycleDaysList,
                                           bad_yearlySchedule)
        bad_setup_data = cycle_calendar_generator.parseScheduleSetup(
                bad_setup_excel)
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_SETUP_FILE,
                cycle_calendar_generator.generateTeacherScheduleCalendar,
                self.schedule_data_good, bad_setup_data)

    def test_raises_valueerror_if_period_end_time_before_start_time(self):
        bad_periodTiming = [
                ["Period Number", "Start Time", "End Time"],
                # Reverse start and end times here
                ["1", datetime.time(9, 0), datetime.time(8, 0)],
                ["2", datetime.time(9, 0), datetime.time(10, 0)],
                ["3", datetime.time(10, 0), datetime.time(11, 0)],
                ["4", datetime.time(11, 0), datetime.time(12, 0)],
                ["5", datetime.time(12, 0), datetime.time(13, 0)]]
        bad_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
        bad_yearlySchedule = [
            ["Date", "Cycle Day"],
            [convert_date_string_to_datetime("August 31, 2018"), bad_cycleDaysList[0]],
            [convert_date_string_to_datetime("September 3, 2018"),
             bad_cycleDaysList[1]],
            [convert_date_string_to_datetime("September 4, 2018"),
             bad_cycleDaysList[2]],
            [convert_date_string_to_datetime("September 5, 2018"),
             bad_cycleDaysList[3]],
            [convert_date_string_to_datetime("September 6, 2018"),
             bad_cycleDaysList[4]],
            [convert_date_string_to_datetime("September 7, 2018"),
             bad_cycleDaysList[5]]]
        bad_setup_excel = make_setup_excel(bad_periodTiming, bad_cycleDaysList,
                                           bad_yearlySchedule)
        bad_setup_data = cycle_calendar_generator.parseScheduleSetup(
                bad_setup_excel)
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_SETUP_FILE,
                cycle_calendar_generator.generateTeacherScheduleCalendar,
                self.schedule_data_good,
                bad_setup_data)

class Test_save_user_schedule_ical(fake_filesystem_unittest.TestCase):
    ## Save Calendar, using filename from user schedule file
    """Tests function to save user schedule ical file"""

    # Test setup
    test_folder_path = "/test-folder"
    output_folder_path = "/test-folder/output"
    user_name = "Test Teacher"
    user_calendar_data = [
            ["20180831 08:00:00", "20180831 09:00:00", "Grade 8"],
            ["20180831 10:00:00", "20180831 11:00:00", "Lunch"],
            ["20180831 11:00:00", "20180831 12:00:00", "Grade 11"],
            ["20180903 09:00:00", "20180903 10:00:00", "Grade 8"],
            ["20180903 10:00:00", "20180903 11:00:00", "Lunch"],
            ["20180903 12:00:00", "20180903 13:00:00", "Grade 11"],
            ["20180904 10:00:00", "20180904 11:00:00", "Lunch"],
            ["20180904 11:00:00", "20180904 12:00:00", "Grade 8"],
            ["20180905 08:00:00", "20180905 09:00:00", "Grade 11"],
            ["20180905 10:00:00", "20180905 11:00:00", "Lunch"],
            ["20180905 12:00:00", "20180905 13:00:00", "Grade 8"],
            ["20180906 09:00:00", "20180906 10:00:00", "Grade 11"],
            ["20180906 10:00:00", "20180906 11:00:00", "Lunch"],
            ["20180907 08:00:00", "20180907 09:00:00", "Grade 8"],
            ["20180907 10:00:00", "20180907 11:00:00", "Lunch"],
            ["20180907 11:00:00", "20180907 12:00:00", "Grade 11"]]
    user_calendar = Calendar()
    for line in user_calendar_data:
        e = Event()
        begin, end, name = line
        begin = arrow.get(begin, "YYYYMMDD HH:mm:ss")
        end = arrow.get(end, "YYYYMMDD HH:mm:ss")
        e.begin = begin
        e.end = end
        e.name = name
        user_calendar.events.add(e)

    def setUp(self):
        self.setUpPyfakefs()
        os.mkdir(self.test_folder_path)
        os.mkdir(self.output_folder_path)

    def test_makes_ical_file_with_right_filename(self):
        saves_correctly = cycle_calendar_generator.saveTeacherScheduleIcal(
                self.user_calendar,
                self.user_name,
                self.output_folder_path)
        self.assertTrue(saves_correctly)
        calendar_filepath = "{}/{}.ics".format(self.output_folder_path,
                                               self.user_name)
        self.assertTrue(os.path.exists(calendar_filepath))

    def test_raises_error_if_incorrect_output_path(self):
        # raises error when given nonexistent output path
        self.assertRaisesRegex(
                ValueError,
                cycle_calendar_generator.ERROR_INVALID_FOLDER,
                cycle_calendar_generator.saveTeacherScheduleIcal,
                self.user_calendar, self.user_name, '/test-notafolder')
