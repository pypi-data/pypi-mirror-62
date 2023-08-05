# -*- coding: utf-8

"""Main module."""
import argparse
import os
import shutil
import datetime
from math import floor
import sys

import openpyxl
from ics import Calendar, Event
import arrow

SCHEDULE_SETUP_FILENAME = 'schedule_setup.xlsx'
OUTPUT_FOLDER_NAME = 'output'

LOCAL_TIMEZONE = arrow.now().datetime.tzinfo
VERSION_MAJOR = sys.version_info.major
VERSION_MINOR = sys.version_info.minor

ERROR_MISSING_SETUP_FILE = 'No schedule setup file found'
ERROR_INVALID_FOLDER = 'Not a valid folder'
ERROR_INVALID_SETUP_FILE = 'Setup file does not follow proper format'
ERROR_SETUP_FILE_NOT_EXCEL = 'Setup file is not a readable Excel file'
ERROR_INVALID_SCHEDULE_FILE = 'Schedule file is not a readable Excel file'
ERROR_NO_TEACHER_FILES = 'Folder has no Excel files for users'
ERROR_ICAL_DIDNT_SAVE = 'A user iCal did not save properly'

SHEET_SETUP_PERIOD_TIMING = 'Period Timing'
SHEET_SETUP_CYCLEDAYSLIST = 'Cycle Days List'
SHEET_SETUP_YEARLYSCHEDULE = 'Yearly Schedule'
SHEET_USER_SCHEDULE = 'User Schedule'

ALL_DONE = 'Schedule iCal generation complete!'


def getArgs():
    return_value = None
    parser = argparse.ArgumentParser(description='Input folder')
    parser.add_argument('folder',
                        nargs='?',
                        action='store',
                        help='folder with input Excel files')
    parsed_args = parser.parse_args()
    if parsed_args.folder is None:
        return_value = os.getcwd()
    else:
        return_value = parsed_args.folder
    return return_value


def readScheduleSetupFile(folder):
    return_value = None
    if (os.path.isdir(folder)):
        schedule_setup_filepath = "{}/{}".format(folder,
                                                 SCHEDULE_SETUP_FILENAME)
        try:
            return_value = openpyxl.load_workbook(schedule_setup_filepath)
        except Exception as e:
            exception_type = str(type(e))
            for case in switch(exception_type):
                if case("<class 'zipfile.BadZipFile'>"):
                    raise ValueError(ERROR_SETUP_FILE_NOT_EXCEL)
                    break
                if case("<class 'FileNotFoundError'>"):
                    raise ValueError(ERROR_MISSING_SETUP_FILE)
                    break
                if case:
                    raise e
    else:
        raise ValueError(ERROR_INVALID_FOLDER)
    return return_value


def parseScheduleSetup(workbook):
    # TODO: Make parser convert non-datetimes into text
    return_value = SetupData()
    try:
        ws_periodTiming = workbook[SHEET_SETUP_PERIOD_TIMING]
        ws_cycleDaysList = workbook[SHEET_SETUP_CYCLEDAYSLIST]
        ws_yearlySchedule = workbook[SHEET_SETUP_YEARLYSCHEDULE]
        rows_periodTiming = tuple(ws_periodTiming.rows)
        rows_periodTiming = rows_periodTiming[1:]
        for row in rows_periodTiming:
            return_value.appendPeriod(row[0].value, row[1].value, row[2].value)
        list_cycleDaysList = []
        for row in ws_cycleDaysList.iter_rows(
                max_row=1, max_col=ws_cycleDaysList.max_column):
            for cell in row:
                list_cycleDaysList.append(cell.value)
        return_value.setCycleDays(list_cycleDaysList)
        rows_yearlySchedule = tuple(ws_yearlySchedule.rows)
        rows_yearlySchedule = rows_yearlySchedule[1:]
        for row in rows_yearlySchedule:
            return_value.appendScheduleDay((row[0].value).date(),
                                           row[1].value)
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case("<class 'KeyError'>"):
                raise ValueError(ERROR_INVALID_SETUP_FILE)
                break
            if case():
                raise e
                break
    return return_value


def readTeacherScheduleFile(filepath):
    return_value = None
    try:
        return_value = openpyxl.load_workbook(filepath)
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case("<class 'zipfile.BadZipFile'>"):
                raise ValueError(ERROR_INVALID_SCHEDULE_FILE)
                break
            if case("<class 'FileNotFoundError'>"):
                raise ValueError(ERROR_INVALID_SCHEDULE_FILE)
                break
            if case:
                raise e
    return return_value


def parseTeacherSchedule(workbook, setupData):
    # TODO: Make parser convert non-datetimes into text
    return_value = ScheduleData(setupData.periodList)
    try:
        ws_userSchedule = workbook[SHEET_USER_SCHEDULE]
        cols_userSchedule = tuple(ws_userSchedule.columns)
        # get period numbers
        schedule_periodNumberCol = cols_userSchedule[0]
        schedule_periodNumberCol = schedule_periodNumberCol[1:]
        setup_periodList = setupData.periodList
        for cell in schedule_periodNumberCol:
            if not (cell.value in setup_periodList):
                raise ValueError(
                    '{}: {}'.format(
                        ERROR_INVALID_SCHEDULE_FILE,
                        'Check that period numbers are the same ' +
                        'in user schedule and setup file'))
        # cycle through columns and add schedule days
        schedule_dayCols = cols_userSchedule[1:]
        for day in schedule_dayCols:
            if day[0].value in setupData.cycleDaysList:
                # turn list of cell objects to list of values
                # TODO: refactor into cell_array_to_values_list function
                day_list = []
                for element in day:
                    day_list.append(element.value)
                return_value.addScheduleDay(day_list)
            else:
                raise ValueError(ERROR_INVALID_SCHEDULE_FILE)
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case("<class 'KeyError'>"):
                raise ValueError(ERROR_INVALID_SCHEDULE_FILE)
                break
            if case():
                raise e
    return return_value


def generateTeacherScheduleCalendar(schedule, setupData):
    return_value = Calendar()
    # For each day in the yearly schedule
    # Get the date and schedule day
    try:
        for date_key in setupData.yearlySchedule:
            # Look up user's schedule for that schedule day
            schedule_day = setupData.yearlySchedule[date_key]
            todays_schedule = schedule.userSchedule[schedule_day]
            # For each entry in that schedule
            for period_number in todays_schedule:
                period_name = todays_schedule[period_number]
                # Look up the start and end times matching the period number
                if period_name:
                    timing = setupData.periodTiming[period_number]
                    start_time = timing[0].replace(tzinfo=LOCAL_TIMEZONE)
                    start = datetime.datetime.combine(date_key, start_time)
                    end_time = timing[1].replace(tzinfo=LOCAL_TIMEZONE)
                    end = datetime.datetime.combine(date_key, end_time)
                    # Use start time, end time and schedule text to make Event
                    e = Event()
                    e.begin = start
                    e.end = end
                    e.name = period_name
                    # Add to calendar
                    return_value.events.add(e)
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case("<class 'KeyError'>"):
                raise ValueError(ERROR_INVALID_SETUP_FILE)
            if case("<class 'ValueError'>"):
                raise ValueError(ERROR_INVALID_SETUP_FILE)
            if case():
                raise e
    return return_value


def saveTeacherScheduleIcal(user_calendar, user_name, folder_path):
    return_value = False
    # Check if folder exists
    try:
        if (os.path.exists(folder_path)):
            user_filepath = "{}/{}.ics".format(folder_path, user_name)
            with open(user_filepath, 'w+') as f:
                f.writelines(user_calendar)
            return_value = True
        else:
            raise ValueError(ERROR_INVALID_FOLDER)
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case():
                raise e
    return return_value


def userScheduleFileScanner(setup_data, folder):
    return_value = False
    try:
        # Check for output sub-folder in given folder
        output_folder_path = os.path.join(folder, OUTPUT_FOLDER_NAME)
        if(os.path.isdir(output_folder_path)):
            # If it exists, delete it and all files inside
            shutil.rmtree(output_folder_path)
        # Make the output sub-folder
        os.mkdir(output_folder_path)
        # Make list of all files in the given folder
        # Remove setup file from list
        folder_contents = []
        folder_scan_results = scandir_with_version_check(folder, VERSION_MAJOR,
                                                         VERSION_MINOR)
        for entry in folder_scan_results:
            if (entry.name != SCHEDULE_SETUP_FILENAME and
                    entry.is_file() and
                    (not entry.name.startswith('.')) and
                    (not entry.name.startswith('~$'))):
                folder_contents.append(entry.path)
        if (len(folder_contents) <= 0):
            raise ValueError(ERROR_NO_TEACHER_FILES)
        # For each remaining file:
        for entry in folder_contents:
            # read filename and remove extension
            try:
                path, filename = os.path.split(entry)
                user_name = os.path.splitext(filename)[0]
                # run readTeacher..., parseTeacher..., etc.
                workbook = readTeacherScheduleFile(entry)
                schedule = parseTeacherSchedule(workbook, setup_data)
                user_calendar = generateTeacherScheduleCalendar(schedule,
                                                                setup_data)
                # When saving, use read filename to name new file
                if (not saveTeacherScheduleIcal(user_calendar, user_name,
                                                output_folder_path)):
                    raise RuntimeError("{}: {}".format(ERROR_ICAL_DIDNT_SAVE,
                                                       user_name))
            except Exception as e:
                exception_type = str(type(e))
                for case in switch(exception_type):
                    if case():
                        raise type(e)('{} - {}'.format(user_name,
                                                       str(e))) from e
        return_value = True
    except Exception as e:
        exception_type = str(type(e))
        for case in switch(exception_type):
            if case():
                raise e
    return return_value


# Convenience objects/functions
class SetupData:
    def __init__(self):
        self.periodList = []
        self.periodTiming = {}
        self.cycleDaysList = []
        self.yearlySchedule = {}

    def appendPeriod(self, periodNumber, startTime, endTime):
        if periodNumber not in self.periodTiming:
            # startTime_parsed = convert_day_fraction_to_time(startTime)
            # endTime_parsed = convert_day_fraction_to_time(endTime)
            self.periodTiming[periodNumber] = (startTime, endTime)
            self.periodList.append(periodNumber)
        else:
            raise ValueError(ERROR_INVALID_SETUP_FILE)

    def setCycleDays(self, cycleDays):
        self.cycleDaysList.clear()
        self.cycleDaysList = cycleDays

    def appendScheduleDay(self, calendar_date, cycleDay):
        if calendar_date not in self.yearlySchedule:
            self.yearlySchedule[calendar_date] = cycleDay
        else:
            raise ValueError(ERROR_INVALID_SETUP_FILE)


class ScheduleData:
    def __init__(self, periodList):
        self.userSchedule = {}
        self.periodList = periodList

    def addScheduleDay(self, schedule_list):
        schedule_day_key = schedule_list[0]
        schedule_periods_list = schedule_list[1:]
        if len(schedule_periods_list) != len(self.periodList):
            raise ValueError(ERROR_INVALID_SCHEDULE_FILE)
        schedule_day = dict(zip(self.periodList, schedule_periods_list))
        self.userSchedule[schedule_day_key] = schedule_day


class switch(object):
    def __init__(self, value):
        self.value = value
        self.fall = False

    def __iter__(self):
        """Return the match method once, then stop"""
        yield self.match
        raise StopIteration

    def match(self, *args):
        """Indicate whether or not to enter a case suite"""
        if self.fall or not args:
            return True
        elif self.value in args:
            self.fall = True
            return True
        else:
            return False


def convert_day_fraction_to_time(day_fraction):
    secs_in_day = datetime.timedelta(days=1).total_seconds()
    total_s = floor(day_fraction*secs_in_day)
    return datetime.time(total_s//3600, (total_s % 3600)//60)


def scandir_with_version_check(folder, version_major, version_minor):
    return_value = []
    if (version_major == 3 and version_minor >= 6):
        with os.scandir(folder) as it:
            for entry in it:
                return_value.append(entry)
    elif (version_major == 3 and version_minor <= 5):
        for entry in os.scandir(folder):
            return_value.append(entry)
    else:
        raise RuntimeError('Please use Python 3.5 or higher')
    return return_value
